#%%
# セットアップ
import numpy as np
import tensorflow as tf
from tensorflow import keras
from tensorflow.keras import layers
from tensorflow.keras import backend as K
import matplotlib.pyplot as plt
from tensorflow.python.keras import activations
from sklearn.preprocessing import StandardScaler

# %%
if __name__ == '__main__':

  # GPU memory を使いすぎない設定 (一部のマシンではこれをいれないでKeras使うとエラーになることもある)
  physical_devices = tf.config.experimental.list_physical_devices('GPU')
  if len(physical_devices) > 0:
    for k in range(len(physical_devices)):
      tf.config.experimental.set_memory_growth(physical_devices[k], True)
      print('memory growth:', tf.config.experimental.get_memory_growth(
          physical_devices[k]))
  else:
    print("Not enough GPU hardware devices available")

#%%
# 定数の設定
batch_size = 64
# Each MNIST image batch is a tensor of shape (batch_size, 28, 28).
# Each input sequence will be of size (28, 28) (height is treated as time).
input_dim = 28

units = 64
output_size = 10  # labels are from 0 to 9

#%%  MNIST データセットの準備

mnist = keras.datasets.mnist

(u_train, y_train), (u_test, y_test) = mnist.load_data()
u_train, u_test = u_train / 255.0, u_test / 255.0
sample, sample_label = u_train[0], y_train[0]

print(f"u_train's shape: {u_train.shape}")
print(f"u_test's shape: {u_test.shape}")

# %% cell をカスタマイズ

## TODO: cell をカスタマイズする．
## TODO: 連続的にcontinuous-sequential-mnist-課題? CSM  
class LeakyRNNCell(tf.keras.layers.SimpleRNNCell):
  # """Cell class for SimpleRNN.

  # See [the Keras RNN API guide](https://www.tensorflow.org/guide/keras/rnn)
  # for details about the usage of RNN API.

  # This class processes one step within the whole time sequence input, whereas
  # `tf.keras.layer.SimpleRNN` processes the whole sequence.

  # Args:
  #   units: Positive integer, dimensionality of the output space.
  #   activation: Activation function to use.
  #     Default: hyperbolic tangent (`tanh`).
  #     If you pass `None`, no activation is applied
  #     (ie. "linear" activation: `a(x) = x`).
  #   use_bias: Boolean, (default `True`), whether the layer uses a bias vector.
  #   kernel_initializer: Initializer for the `kernel` weights matrix,
  #     used for the linear transformation of the inputs. Default:
  #     `glorot_uniform`.
  #   recurrent_initializer: Initializer for the `recurrent_kernel`
  #     weights matrix, used for the linear transformation of the recurrent state.
  #     Default: `orthogonal`.
  #   bias_initializer: Initializer for the bias vector. Default: `zeros`.
  #   kernel_regularizer: Regularizer function applied to the `kernel` weights
  #     matrix. Default: `None`.
  #   recurrent_regularizer: Regularizer function applied to the
  #     `recurrent_kernel` weights matrix. Default: `None`.
  #   bias_regularizer: Regularizer function applied to the bias vector. Default:
  #     `None`.
  #   kernel_constraint: Constraint function applied to the `kernel` weights
  #     matrix. Default: `None`.
  #   recurrent_constraint: Constraint function applied to the `recurrent_kernel`
  #     weights matrix. Default: `None`.
  #   bias_constraint: Constraint function applied to the bias vector. Default:
  #     `None`.
  #   dropout: Float between 0 and 1. Fraction of the units to drop for the linear
  #     transformation of the inputs. Default: 0.
  #   recurrent_dropout: Float between 0 and 1. Fraction of the units to drop for
  #     the linear transformation of the recurrent state. Default: 0.

  # Call arguments:
  #   inputs: A 2D tensor, with shape of `[batch, feature]`.
  #   states: A 2D tensor with shape of `[batch, units]`, which is the state from
  #     the previous time step. For timestep 0, the initial state provided by user
  #     will be feed to cell.
  #   training: Python boolean indicating whether the layer should behave in
  #     training mode or in inference mode. Only relevant when `dropout` or
  #     `recurrent_dropout` is used.

  # Examples:

  # ```python
  # inputs = np.random.random([32, 10, 8]).astype(np.float32)
  # rnn = tf.keras.layers.RNN(tf.keras.layers.SimpleRNNCell(4))

  # output = rnn(inputs)  # The output has shape `[32, 4]`.

  # rnn = tf.keras.layers.RNN(
  #     tf.keras.layers.SimpleRNNCell(4),
  #     return_sequences=True,
  #     return_state=True)

  # # whole_sequence_output has shape `[32, 10, 4]`.
  # # final_state has shape `[32, 4]`.
  # whole_sequence_output, final_state = rnn(inputs)
  # ```
  # """

  def __init__(self,
               units,
               alpha=0.1, # add
               activation='tanh',
               use_bias=True,
               kernel_initializer='glorot_uniform',
               recurrent_initializer='orthogonal',
               bias_initializer='zeros',
               kernel_regularizer=None,
               recurrent_regularizer=None,
               bias_regularizer=None,
               kernel_constraint=None,
               recurrent_constraint=None,
               bias_constraint=None,
               dropout=0.,
               recurrent_dropout=0.,
               **kwargs):

    super(LeakyRNNCell, self).__init__( 
               units=units,
               activation=activation,
               use_bias=use_bias,
               kernel_initializer=kernel_initializer,
               recurrent_initializer=recurrent_initializer,
               bias_initializer=bias_initializer,
               kernel_regularizer=kernel_regularizer,
               recurrent_regularizer=recurrent_regularizer,
               bias_regularizer=bias_regularizer,
               kernel_constraint=kernel_constraint,
               recurrent_constraint=recurrent_constraint,
               bias_constraint=bias_constraint,
               dropout=dropout,
               recurrent_dropout=recurrent_dropout,
               **kwargs)
    
    self.alpha = alpha


  def build(self, input_shape):
    super().build(input_shape)

  # 以下のcallを変更することでモデルをカスタマイズできる
  def call(self, inputs, states, training=None):
    prev_output = states[0] if tf.nest.is_nested(states) else states
    dp_mask = self.get_dropout_mask_for_cell(inputs, training)
    rec_dp_mask = self.get_recurrent_dropout_mask_for_cell(
        prev_output, training)

    # backend.dot は tf.matmul に変更した．

    # ここは外部入力 に入力の重みを掛けている部分
    if dp_mask is not None:
      h = tf.matmul(inputs * dp_mask, self.kernel)
    else:
      h = tf.matmul(inputs, self.kernel)
    if self.bias is not None:
      h = tf.add(h, self.bias)

    # ここは内部状態に再帰結合行列を掛けている部分
    if rec_dp_mask is not None:
      prev_output = prev_output * rec_dp_mask
    
    # ここを変更 (1-alpha)x_{t-1} + alpha * (wx + w u)
    new_state = (1-self.alpha)* prev_output + self.alpha * (
               h + tf.matmul(self.activation(prev_output), self.recurrent_kernel))
    
    if self.activation is not None:
      output = self.activation(new_state)
    else:
      output = new_state
    
    new_state = [output] if tf.nest.is_nested(states) else output

    return output, new_state

#%%
rnn_cell = LeakyRNNCell(units=100, alpha=0.1)

#%%

class LeakyRNNModel(tf.keras.Model):   # tf.keras.Model クラスを継承してクラスを作る

  def __init__(self, N, N_in, N_out, alpha=0.1, **kwargs):
    '''
    N: ユニット数
    N_in: 入力ベクトルの次元
    N_out: 出力の次元
    '''
    super(LeakyRNNModel, self).__init__(**kwargs)  # 元のクラスのinitを実行

    # 自分のメンバ関数に情報を登録する
    self.N = N
    self.N_in = N_in
    self.N_out = N_out

    #各レイヤーを作る．
    self.rnn_cell = LeakyRNNCell(self.N, alpha=alpha)
    self.rnn_layer = keras.layers.RNN(self.rnn_cell, 
                                      input_shape=(None, self.N_in),
                                      return_sequences=True)    # RNNレイヤー作る． return_sequencesを使う．
    self.dense = keras.layers.Dense(self.N_out)

  # 入力を受け取って出力を計算する流れを書く関数
  def call(self, inputs, training=False):  
    ''' inputs: shape is [bs, length_of_timestep, N_in]  '''
    self.x = self.rnn_layer(inputs, training=training) # self.x に入れておくことによって，後から値を参照できる.
    return self.dense(self.x[:,-1]) # 最後の部分だけを使う．

def build_leaky_rnn_model(N, N_in=28, N_out=10, alpha=0.1):
  model = LeakyRNNModel(N=N, N_in=N_in, N_out=N_out, alpha=alpha)

  inputs = tf.random.uniform(shape=[batch_size,100,28]) # 入力を入れたときにパラメータのサイズが確定し，重みの初期化が行われる．
  model(inputs)

  # わかりやすく呼び出しやすい別名を付けておく
  model.W_rec = model.rnn_cell.recurrent_kernel
  model.W_in = model.rnn_cell.kernel
  model.bias = model.rnn_cell.bias
  model.W_out = model.dense.kernel
  model.b_out = model.dense.bias

  return model

model = build_leaky_rnn_model(N=100, N_in=28, N_out=10, alpha=0.1)

print(f'W_rec.shape= {model.W_rec.shape}')  # [100, 100]
print(f'W_in.shape= {model.W_in.shape}')    # [28, 100]
print(f'W_out.shape= {model.W_out.shape}')  # [100, 10]

# %% 入力を入れてみるテスト
z = model(u_train[0:5])  # 5バッチ分のシーケンスをいれてみる．
print(z)  # 出力サイズは(5,10)
# %%
# リッジ回帰のロスクラス．
class RidgeLoss:
      
      def __init__(self, alpha):
          self.alpha = alpha
          self.E = None
      
      def fit(self, d, y, W):
          d_y = tf.norm(d - y) * 0.5
          alpha_W = tf.norm(W) * self.alpha * 0.5
          self.E = tf.add(d_y, alpha_W)

# %%
# 誤差を評価する関数 from_logits をTrueにしておくと，評価時にsoftmaxを使ってくれる
cross_entropy = tf.keras.losses.SparseCategoricalCrossentropy(
    from_logits=True)
# 学習法
optimizer = tf.keras.optimizers.SGD(lr=0.005)
# メトリック (誤差や正解率を記録するもの)
metrics_train_loss = tf.keras.metrics.Mean()
metrics_train_accuracy = tf.keras.metrics.SparseCategoricalAccuracy()
metrics_test_loss = tf.keras.metrics.Mean()
metrics_test_accuracy = tf.keras.metrics.SparseCategoricalAccuracy()

# %%

# Hebb学習則により、重みの変化量を返す関数．
def hebb(x, y):
  
  # x size (batch_size, 28, 100), y size (batch_size, 28, 100)
  dW = tf.matmul(x, y, transpose_a=True)  # (batch_size, 100, 100)
  
  return dW

# リッジ回帰を行うクラス
class RidgeReg:

    def __init__(self, alpha):
        self.alpha = alpha
        self.W = None

    def fit(self, x, d):
        # x.T * x の値  (100,data_size) * (data_size,100) → (100,100)
        x_x = tf.matmul(x, x, transpose_a=True)
        # 単位行列を作成  (100,100)
        i = tf.eye(x_x.shape[1])
        # x.T * y の値  (100,data_size) * (data_size, 10) → (100, 10)
        x_d = tf.matmul(x, d, transpose_a=True)

        # 損失関数を最小化した重みを求める式  (100,100) * (100,10) → (100, 10)
        self.W = tf.matmul(tf.linalg.inv(x_x + self.alpha * i), x_d)

# %%

def train_step_hebb(inputs):
    #
    #  inputs: shape=(batch_size, length_of_timestep, dim) = (?, 28, 28)
    #
    
    # with tf.GradientTape() as tape:
    model(inputs)
    
    # dW size (batch_size, 100, 100)
    dW = hebb(model.x, K.sigmoid(model.x))
    # dW_sum size (100,100), 0.0000001:学習率
    dW_sum = tf.reduce_sum(dW, 0) * 0.0000001
    # model.W_rec size (100, 100)
    model.W_rec.assign(model.W_rec + dW_sum)
  
#%%
def eva_step_train(inputs,labels):
  
    z = model(inputs)
    #loss = cross_entropy(labels, z)
    
    #metrics_train_loss(loss)
    metrics_train_accuracy(labels, z)
  
def eva_step_test(inputs,labels):
  
    z = model(inputs)
    #loss = cross_entropy(labels, z)
    
    #metrics_test_loss(loss)
    metrics_test_accuracy(labels, z)

# %%

train_step_hebb(u_train[0:5])
train_step_hebb(u_test[0:5])

eva_step_train(u_train[0:5], y_train[0:5])
eva_step_test(u_test[0:5], y_test[0:5])

# %% 訓練データ、テストデータをdataset化

ds_train = tf.data.Dataset.from_tensor_slices((u_train, y_train)).shuffle(
    buffer_size=60000).batch(batch_size).prefetch(10000)

ds_test = tf.data.Dataset.from_tensor_slices((u_test, y_test)).shuffle(
    buffer_size=10000).batch(batch_size).prefetch(10000)

# %% 
# dataset の動きの確認　(訓練データ)
for x, y in ds_train.take(1):
  print("[train_data]")
  print(x.shape)
  print(y.shape)
  print(y)

print()

# dataset の動きの確認　(テストデータ)
for x, y in ds_test.take(1):
  print("[test_data]")
  print(x.shape)
  print(y.shape)
  print(y)

# %%

def train(epochs):

  train_losses = []  # 記録用のリスト
  train_accs = []
  test_losses = []
  test_accs = []
  epochs_list = []

  for ep in range(epochs):
    
    epochs_list.append(ep+1)
    
    y = None
    x = None
    d = None

    # hebb学習則による学習
    for input_seq, label in ds_train:
      train_step_hebb(input_seq)
    
    # リッジ回帰のデータ収集
    for input_seq, label in ds_train:
      if y == None:
        y = model(input_seq)
      else:
        y = tf.concat([y, model(input_seq)], 0)
      
      if x == None:
        # x size (batch_size, 100)
        x = model.x[:,-1]
      else:
        # x size (data_size, 100)
        x = tf.concat([x, model.x[:,-1]], 0)
      
      if d == None:
        # d size (batch_size)
        d = label
      else:
        # d size (data_size)
        d = tf.concat([d, label], 0)
    
    # リッジ回帰による重みの更新
    #x = tf.cast(StandardScaler().fit_transform(x), tf.float32)
    d = tf.one_hot(d, depth=10)   # d size (data_size, 10)
    reg = RidgeReg(alpha = 0.001)
    reg.fit(x, d)
    model.W_out.assign(reg.W)
    
    # リッジ回帰によるロスの計算(訓練)
    y = tf.argmax(y, 1)
    y = tf.one_hot(y, depth=10)   # y size (data_size, 10)
    loss = RidgeLoss(alpha = 0.001)
    loss.fit(d, y, model.W_out)
    metrics_train_loss(loss.E)
    
    # accuracy の収集
    for input_seq, label in ds_train:
      eva_step_train(input_seq, label)
    
    print(f'EP {ep+1} train loss: {metrics_train_loss.result()}, train accuracy: {metrics_train_accuracy.result()}')
    
    train_losses.append(metrics_train_loss.result())
    train_accs.append(metrics_train_accuracy.result())

    metrics_train_loss.reset_states()
    metrics_train_accuracy.reset_states()
    
    y = None
    d = None
    
    for input_seq, label in ds_test:
      if y == None:
        y = model(input_seq)
      else:
        y = tf.concat([y, model(input_seq)], 0)
          
      if d == None:
        # d size (batch_size)
        d = label
      else:
        # d size (data_size)
        d = tf.concat([d, label], 0)
    
    # リッジ回帰によるロスの計算(テスト)
    d = tf.one_hot(d, depth=10)   # d size (data_size, 10)
    y = tf.argmax(y, 1)
    y = tf.one_hot(y, depth=10)   # y size (data_size, 10)
    loss = RidgeLoss(alpha = 0.001)
    loss.fit(d, y, model.W_out)
    metrics_test_loss(loss.E)
    
    # accuracy の収集
    for input_seq, label in ds_test:
      eva_step_test(input_seq, label)

    print(f'EP {ep+1} test loss: {metrics_test_loss.result()}, test accuracy: {metrics_test_accuracy.result()}')

    test_losses.append(metrics_test_loss.result())
    test_accs.append(metrics_test_accuracy.result())

    metrics_test_loss.reset_states()
    metrics_test_accuracy.reset_states()

  return train_losses, train_accs, test_losses, test_accs, epochs_list

# %%
epochs = 5
train_losses, train_accs, test_losses, test_accs, epochs_list = train(epochs)
print('training & testing was finished.')

# %%
plt.plot(train_losses)
# %%
plt.plot(train_accs)
# %%
plt.plot(test_losses)
# %%
plt.plot(test_accs)
# %%
# 結果をexcelに出力
import openpyxl

#  新しいブックを作成
wb = openpyxl.Workbook()

# シートのオブジェクトを取り出す
s1 = wb.get_sheet_by_name(wb.get_sheet_names()[0])

# convertするためにtensorをnumpyのlistに変換
train_accs_list = np.array(train_accs)
test_accs_list = np.array(test_accs)

# EXCELに書き込み
for i in range(epochs):
    # A列
    s1.cell(row=i+1,column=1,value=epochs_list[i])
    # B列
    s1.cell(row=i+1,column=2,value=train_accs_list[i])
    # C列
    s1.cell(row=i+1,column=3,value=test_accs_list[i])
    
wb.save('hebb.xlsx')

# %%
